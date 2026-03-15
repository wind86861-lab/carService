import io
from datetime import datetime


def _fmt(n) -> str:
    try:
        return f"{float(n):,.0f}"
    except (TypeError, ValueError):
        return "0"


def generate_excel_report(orders: list, summary: dict) -> bytes:
    """Build an .xlsx report in memory and return raw bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()

    ws_sum = wb.active
    ws_sum.title = "Summary"
    header_font = Font(bold=True)
    labels = [
        ("Total Revenue (UZS)", summary.get("total_revenue", 0)),
        ("Total Parts Cost (UZS)", summary.get("total_parts_cost", 0)),
        ("Total Profit (UZS)", summary.get("total_profit", 0)),
        ("Service Share 60% (UZS)", summary.get("total_service_share", 0)),
        ("Masters Share 40% (UZS)", summary.get("total_master_share", 0)),
    ]
    for row_idx, (label, value) in enumerate(labels, start=1):
        ws_sum.cell(row=row_idx, column=1, value=label).font = header_font
        cell = ws_sum.cell(row=row_idx, column=2, value=float(value))
        cell.alignment = Alignment(horizontal="right")
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 20

    ws_ord = wb.create_sheet("Orders")
    col_headers = [
        "Order #", "Master", "Car", "Plate",
        "Revenue", "Parts Cost", "Profit", "Master Share", "Service Share", "Closed At",
    ]
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for col_idx, header in enumerate(col_headers, start=1):
        cell = ws_ord.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for row_idx, o in enumerate(orders, start=2):
        brand = o.get("brand") or ""
        model = o.get("model") or ""
        car = f"{brand} {model}".strip() or "—"
        closed_at = o.get("closed_at")
        closed_str = closed_at.strftime("%d.%m.%Y") if hasattr(closed_at, "strftime") else str(closed_at or "")
        values = [
            o.get("order_number", ""),
            o.get("master_name", ""),
            car,
            o.get("plate", ""),
            float(o.get("agreed_price", 0)),
            float(o.get("parts_cost", 0)),
            float(o.get("profit", 0)),
            float(o.get("master_share", 0)),
            float(o.get("service_share", 0)),
            closed_str,
        ]
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, val in enumerate(values, start=1):
            cell = ws_ord.cell(row=row_idx, column=col_idx, value=val)
            if isinstance(val, float):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0'
            if fill:
                cell.fill = fill

    for col_idx in range(1, len(col_headers) + 1):
        ws_ord.column_dimensions[ws_ord.cell(row=1, column=col_idx).column_letter].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_pdf_report(orders: list, summary: dict, period_label: str = "") -> bytes:
    """Build a .pdf report in memory and return raw bytes."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=1.5 * cm, rightMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("AutoService Financial Report", styles["Title"]))
    if period_label:
        elements.append(Paragraph(f"Period: {period_label}", styles["Normal"]))
    elements.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]))
    elements.append(Spacer(1, 0.4 * cm))

    sum_data = [
        ["Metric", "Value (UZS)"],
        ["Total Revenue", _fmt(summary.get("total_revenue", 0))],
        ["Total Parts Cost", _fmt(summary.get("total_parts_cost", 0))],
        ["Total Profit", _fmt(summary.get("total_profit", 0))],
        ["Service Share (60%)", _fmt(summary.get("total_service_share", 0))],
        ["Masters Share (40%)", _fmt(summary.get("total_master_share", 0))],
    ]
    sum_table = Table(sum_data, colWidths=[8 * cm, 6 * cm])
    sum_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
    ]))
    elements.append(sum_table)
    elements.append(Spacer(1, 0.6 * cm))

    if orders:
        elements.append(Paragraph("Order Breakdown", styles["Heading2"]))
        elements.append(Spacer(1, 0.2 * cm))
        ord_data = [["Order #", "Master", "Car", "Revenue", "Parts", "Profit", "M.Share", "S.Share", "Closed"]]
        for o in orders:
            brand = o.get("brand") or ""
            model = o.get("model") or ""
            car = f"{brand} {model}".strip() or "—"
            closed_at = o.get("closed_at")
            closed_str = closed_at.strftime("%d.%m.%Y") if hasattr(closed_at, "strftime") else str(closed_at or "")
            ord_data.append([
                o.get("order_number", ""),
                o.get("master_name", ""),
                car,
                _fmt(o.get("agreed_price", 0)),
                _fmt(o.get("parts_cost", 0)),
                _fmt(o.get("profit", 0)),
                _fmt(o.get("master_share", 0)),
                _fmt(o.get("service_share", 0)),
                closed_str,
            ])
        ord_table = Table(ord_data, repeatRows=1)
        ord_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (3, 0), (-2, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
        ]))
        elements.append(ord_table)

    doc.build(elements)
    return buf.getvalue()
